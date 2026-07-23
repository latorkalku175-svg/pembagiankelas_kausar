from django.contrib import messages
import json as _json_module
import os
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.db.models import Avg
from django.http import HttpResponse, JsonResponse
from django.views.decorators.cache import cache_control
from django.templatetags.static import static
from django.utils import timezone

from .models import (
    Kelas, Siswa, MateriUpload, ProfilGuru, RiwayatUploadXlsx, RiwayatPerbangku,
    RiwayatClusterKmeans,
)
from .clustering import run_clustering, _kumpulkan_kelas_setingkat
from .excel_import import proses_upload_dan_cluster, ImportError_
from .promosi import get_kelas_per_tingkat, proses_naik_kelas
from .perbangku import atur_meja_per_kelas
from .pdf_perbangku import build_perbangku_pdf
from .pdf_cluster import build_cluster_pdf
from .pdf_data_olahan import build_data_olahan_pdf
from .forms import (
    LoginForm, KelasForm, SiswaForm, SiswaManualForm, ProfilGuruForm,
    UploadDataSiswaForm, INPUT_CLASS,
)


def _style_password_form(form):
    for field in form.fields.values():
        field.widget.attrs.update({"class": INPUT_CLASS})
    return form


# ---------------------------------------------------------------------------
# Helper: kelas yang sedang aktif (dipilih lewat dropdown di header)
# ---------------------------------------------------------------------------
def _get_kelas_aktif(request):
    kelas_list = Kelas.objects.filter(guru=request.user)
    if not kelas_list.exists():
        return None, kelas_list

    kelas_id = request.session.get("kelas_aktif_id")
    kelas = kelas_list.filter(id=kelas_id).first()
    if kelas is None:
        kelas = kelas_list.first()
        request.session["kelas_aktif_id"] = kelas.id
    return kelas, kelas_list


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------
def login_view(request):
    if request.user.is_authenticated:
        return redirect("dashboard:beranda")

    form = LoginForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = authenticate(
            request,
            username=form.cleaned_data["username"],
            password=form.cleaned_data["password"],
        )
        if user is not None:
            login(request, user)
            return redirect("dashboard:beranda")
        messages.error(request, "Username atau password salah.")
    return render(request, "dashboard/login.html", {"form": form})


def logout_view(request):
    logout(request)
    return redirect("dashboard:login")


# ---------------------------------------------------------------------------
# Beranda
# ---------------------------------------------------------------------------
@login_required
def beranda(request):
    kelas, kelas_list = _get_kelas_aktif(request)

    cluster_counts = {"cerdas": 0, "pintar": 0, "malas": 0}
    if kelas:
        for label in cluster_counts:
            cluster_counts[label] = kelas.siswa_list.filter(cluster_label=label).count()

    context = {
        "kelas": kelas,
        "kelas_list": kelas_list,
        "cluster_counts": cluster_counts,
        "active_nav": "beranda",
    }
    return render(request, "dashboard/beranda.html", context)


@login_required
def pilih_kelas(request, kelas_id):
    kelas = get_object_or_404(Kelas, id=kelas_id, guru=request.user)
    request.session["kelas_aktif_id"] = kelas.id
    next_url = request.GET.get("next") or reverse("dashboard:beranda")
    return redirect(next_url)


# ---------------------------------------------------------------------------
# Kelas (manajemen kelas & siswa)
# ---------------------------------------------------------------------------
@login_required
def kelas_list_view(request):
    kelas_list = Kelas.objects.filter(guru=request.user)

    if request.method == "POST":
        form = KelasForm(request.POST)
        if form.is_valid():
            baru = form.save(commit=False)
            baru.guru = request.user
            baru.save()
            messages.success(request, f"Kelas '{baru.nama}' berhasil ditambahkan.")
            return redirect("dashboard:kelas_detail", kelas_id=baru.id)
    else:
        form = KelasForm()

    kelas, _ = _get_kelas_aktif(request)
    context = {
        "kelas_list": kelas_list,
        "kelas": kelas,
        "form": form,
        "active_nav": "kelas",
    }
    return render(request, "dashboard/kelas_list.html", context)


@login_required
def kelas_detail_view(request, kelas_id):
    kelas_obj = get_object_or_404(Kelas, id=kelas_id, guru=request.user)
    kelas, kelas_list = _get_kelas_aktif(request)

    context = {
        "kelas_obj": kelas_obj,
        "siswa_list": kelas_obj.siswa_list.all(),
        "kelas": kelas,
        "kelas_list": kelas_list,
        "active_nav": "kelas",
    }
    return render(request, "dashboard/kelas_detail.html", context)


@login_required
def kelas_hapus_view(request, kelas_id):
    kelas_obj = get_object_or_404(Kelas, id=kelas_id, guru=request.user)
    nama = kelas_obj.nama
    kelas_obj.delete()
    messages.success(request, f"Kelas '{nama}' dihapus.")
    if request.session.get("kelas_aktif_id") == kelas_id:
        request.session.pop("kelas_aktif_id", None)
    return redirect("dashboard:kelas_list")


@login_required
def siswa_edit_view(request, kelas_id, siswa_id):
    kelas_obj = get_object_or_404(Kelas, id=kelas_id, guru=request.user)
    siswa = get_object_or_404(Siswa, id=siswa_id, kelas=kelas_obj)

    if request.method == "POST":
        form = SiswaForm(request.POST, instance=siswa)
        if form.is_valid():
            form.save()
            messages.success(request, f"Data '{siswa.nama}' diperbarui.")
            return redirect("dashboard:kelas_detail", kelas_id=kelas_obj.id)
    else:
        form = SiswaForm(instance=siswa)

    kelas, kelas_list = _get_kelas_aktif(request)
    context = {
        "kelas_obj": kelas_obj,
        "siswa": siswa,
        "form": form,
        "kelas": kelas,
        "kelas_list": kelas_list,
        "active_nav": "kelas",
    }
    return render(request, "dashboard/siswa_edit.html", context)


@login_required
def siswa_hapus_view(request, kelas_id, siswa_id):
    kelas_obj = get_object_or_404(Kelas, id=kelas_id, guru=request.user)
    siswa = get_object_or_404(Siswa, id=siswa_id, kelas=kelas_obj)
    siswa.delete()
    messages.success(request, f"Siswa '{siswa.nama}' dihapus.")
    return redirect("dashboard:kelas_detail", kelas_id=kelas_obj.id)


# ---------------------------------------------------------------------------
# Cluster (K-Means)
# ---------------------------------------------------------------------------
def _judul_gabungan_kelas(kelompok_kelas):
    """
    Bikin judul gabungan dari daftar Kelas setingkat, mis. ["Kelas 1A", "Kelas
    1B"] -> "Kelas 1A dan 1B", atau 3+ kelas -> "Kelas 1A, 1B, dan 1C".
    Kalau cuma 1 kelas, kembalikan nama aslinya apa adanya (tidak digabung).
    """
    if len(kelompok_kelas) <= 1:
        return kelompok_kelas[0].nama if kelompok_kelas else ""

    nama_urut = sorted(k.nama for k in kelompok_kelas)
    pertama, *sisanya = nama_urut
    prefix = "Kelas "
    sisanya_pendek = [n[len(prefix):] if n.startswith(prefix) else n for n in sisanya]

    if len(sisanya_pendek) == 1:
        return f"{pertama} dan {sisanya_pendek[0]}"
    return f"{pertama}, " + ", ".join(sisanya_pendek[:-1]) + f", dan {sisanya_pendek[-1]}"


def _bangun_snapshot_cluster(kelompok_kelas, judul_kelas):
    """Ambil hasil clustering K-Means saat ini untuk seluruh siswa di
    `kelompok_kelas`, lalu bekukan jadi dict yang bisa disimpan sebagai JSON
    (RiwayatClusterKmeans.snapshot) supaya preview/unduh riwayat lama tetap
    menampilkan hasil persis seperti saat itu, walau data siswa
    berubah/di-cluster ulang belakangan. Dipakai bersama oleh PDF
    preview/unduh (state sekarang) dan pencatatan riwayat, supaya logikanya
    selalu konsisten dengan math_perbangku._bangun_snapshot_perbangku.

    Kembalikan (snapshot, ringkasan) -- ringkasan berisi jumlah
    cerdas/pintar/malas untuk dipakai mengisi kolom RiwayatClusterKmeans.
    """
    siswa_qs = (
        Siswa.objects.filter(kelas__in=kelompok_kelas)
        .select_related("kelas")
        .order_by("kelas__nama", "rangking", "nama")
    )

    urutan_label = {"cerdas": 0, "pintar": 1, "malas": 2}
    ringkasan = {"cerdas": 0, "pintar": 0, "malas": 0}
    siswa_snapshot = []
    for s in siswa_qs:
        if s.cluster_label in ringkasan:
            ringkasan[s.cluster_label] += 1
        siswa_snapshot.append({
            "nama": s.nama,
            "kelas_nama": s.kelas.nama,
            "cluster_label": s.cluster_label,
            "eksakta": s.eksakta if s.eksakta is not None else s.nilai_matematika,
            "nonekstakta": s.nonekstakta if s.nonekstakta is not None else s.nilai_bahasa,
            "rangking": s.rangking,
            "izin": s.izin,
            "sakit": s.sakit,
            "alpa": s.alpa,
            "skor_akademik": s.skor_akademik,
            "skor_disiplin": s.skor_disiplin,
        })

    siswa_snapshot.sort(key=lambda x: (
        urutan_label.get(x["cluster_label"], 3),
        -(x["skor_akademik"]) if x["skor_akademik"] is not None else 0,
    ))

    snapshot = {
        "judul_kelas": judul_kelas,
        "kelompok_kelas": [k.nama for k in kelompok_kelas],
        "siswa": siswa_snapshot,
    }
    return snapshot, ringkasan


def _catat_riwayat_cluster(kelas, kelompok_kelas, judul_kelas, user, sumber):
    """Simpan snapshot hasil clustering K-Means saat ini sebagai satu baris
    RiwayatClusterKmeans, dipakai baik oleh 'Jalankan Ulang K-Means' manual
    maupun oleh upload Excel leger (proses_upload_dan_cluster)."""
    snapshot, ringkasan = _bangun_snapshot_cluster(kelompok_kelas, judul_kelas)
    RiwayatClusterKmeans.objects.create(
        kelas=kelas,
        dijalankan_oleh=user,
        judul_kelas=judul_kelas,
        jumlah_siswa=len(snapshot["siswa"]),
        jumlah_cerdas=ringkasan["cerdas"],
        jumlah_pintar=ringkasan["pintar"],
        jumlah_malas=ringkasan["malas"],
        sumber=sumber,
        snapshot=snapshot,
    )


@login_required
def cluster_overview(request):
    kelas, kelas_list = _get_kelas_aktif(request)
    import json as _json

    ringkasan = None
    scatter_data = []
    profil_fitur = {}
    judul_kelas = kelas.nama if kelas else ""
    kelompok_kelas = []

    if kelas:
        # Gabungkan dengan kelas lain yang setingkat (mis. Kelas 1A + Kelas 1B)
        # supaya tampilan overview-nya konsisten dengan cara run_clustering()
        # menghitung -- lihat dashboard/clustering.py::_kumpulkan_kelas_setingkat.
        kelompok_kelas = _kumpulkan_kelas_setingkat(kelas)
        judul_kelas = _judul_gabungan_kelas(kelompok_kelas)

        siswa_gabungan = Siswa.objects.filter(kelas__in=kelompok_kelas)

        ringkasan = {
            "cerdas": siswa_gabungan.filter(cluster_label="cerdas").count(),
            "pintar": siswa_gabungan.filter(cluster_label="pintar").count(),
            "malas": siswa_gabungan.filter(cluster_label="malas").count(),
            "belum": siswa_gabungan.filter(cluster_label__isnull=True).count(),
        }

        # Data scatter: Skor_Akademik vs Skor_Disiplin (Cell 43 notebook)
        for s in siswa_gabungan.exclude(skor_akademik=None):
            scatter_data.append({
                "x": round(s.skor_akademik, 3),
                "y": round(s.skor_disiplin, 3),
                "label": s.cluster_label or "belum",
                "nama": s.nama,
            })

        # Profil rata-rata fitur per kategori (Cell 44 notebook: box plot)
        FITUR = ["eksakta", "nonekstakta", "rangking", "izin", "sakit", "alpa"]
        LABEL_FITUR = ["Eksakta", "NonEksakta", "Rangking", "Izin", "Sakit", "Alpa"]
        for label in ("cerdas", "pintar", "malas"):
            qs = list(siswa_gabungan.filter(cluster_label=label).values(*FITUR))
            for i, fitur in enumerate(FITUR):
                vals = [r[fitur] for r in qs if r[fitur] is not None]
                profil_fitur.setdefault(LABEL_FITUR[i], {})[label] = round(
                    sum(vals) / len(vals), 2
                ) if vals else 0

    # Crosstab semua kelas (Cell 49 notebook)
    semua_kelas_ringkasan = []
    for k in kelas_list:
        row = {
            "kelas": k,
            "jumlah": k.siswa_list.count(),
            "cerdas": k.siswa_list.filter(cluster_label="cerdas").count(),
            "pintar": k.siswa_list.filter(cluster_label="pintar").count(),
            "malas": k.siswa_list.filter(cluster_label="malas").count(),
        }
        semua_kelas_ringkasan.append(row)

    context = {
        "kelas": kelas,
        "kelas_list": kelas_list,
        "judul_kelas": judul_kelas,
        "kelompok_kelas": kelompok_kelas,
        "ringkasan": ringkasan,
        "semua_kelas_ringkasan": semua_kelas_ringkasan,
        "scatter_data_json": _json.dumps(scatter_data),
        "profil_fitur_json": _json.dumps(profil_fitur),
        "active_nav": "cluster",
    }
    return render(request, "dashboard/cluster_overview.html", context)


@login_required
def upload_data_siswa_view(request):
    if request.method == "POST":
        form = UploadDataSiswaForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            guru_tujuan = form.cleaned_data.get("guru_tujuan") if request.user.is_superuser else None
            target_guru = guru_tujuan or request.user
            try:
                ringkasan, peringatan = proses_upload_dan_cluster(
                    form.cleaned_data["files"], target_guru, diupload_oleh=request.user
                )
                for pesan in peringatan:
                    messages.warning(request, pesan)
                nama_kelas_list = ", ".join(ringkasan.keys())
                total_siswa = sum(r["jumlah"] for r in ringkasan.values())
                # Catat satu riwayat cluster K-Means per rombel (kelas) yang
                # baru selesai diproses -- proses_upload_dan_cluster() sudah
                # menggabungkan seluruh file yang diupload bareng untuk
                # perhitungan z-score/KMeans-nya, di sini riwayatnya cukup
                # dipisah per rombel supaya sejalan dengan cara ditampilkan
                # di halaman Cluster.
                for nama_kelas in ringkasan.keys():
                    kelas_obj = Kelas.objects.filter(guru=target_guru, nama=nama_kelas).first()
                    if kelas_obj:
                        _catat_riwayat_cluster(
                            kelas=kelas_obj, kelompok_kelas=[kelas_obj],
                            judul_kelas=nama_kelas, user=request.user,
                            sumber="upload_excel",
                        )
                messages.success(
                    request,
                    f"Berhasil melatih K-Means untuk {total_siswa} siswa dan "
                    f"mengelompokkan ke dalam {nama_kelas_list} "
                    f"(guru: {target_guru.first_name or target_guru.username}).",
                )
                # arahkan kelas aktif ke salah satu kelas yang baru saja diproses
                # (cuma kalau target guru = guru yang sedang login)
                if target_guru == request.user:
                    pertama = Kelas.objects.filter(guru=request.user, nama__in=ringkasan.keys()).first()
                    if pertama:
                        request.session["kelas_aktif_id"] = pertama.id
            except ImportError_ as exc:
                messages.error(request, str(exc))
        else:
            for field_errors in form.errors.values():
                for err in field_errors:
                    messages.error(request, err)
    return redirect("dashboard:upload_file")


@login_required
def cluster_jalankan(request):
    kelas, _ = _get_kelas_aktif(request)
    if kelas:
        if kelas.siswa_list.count() == 0:
            messages.warning(request, "Belum ada data siswa di kelas ini untuk dikelompokkan.")
        else:
            run_clustering(kelas)
            # run_clustering() menggabungkan perhitungan dengan kelas lain yang
            # setingkat (mis. 1A + 1B) kalau ada, supaya konsisten dengan hasil
            # upload Excel -- beri tahu guru kalau itu terjadi.
            kelompok_kelas = _kumpulkan_kelas_setingkat(kelas)
            judul_kelas = _judul_gabungan_kelas(kelompok_kelas)
            _catat_riwayat_cluster(
                kelas=kelas, kelompok_kelas=kelompok_kelas, judul_kelas=judul_kelas,
                user=request.user, sumber="manual",
            )
            kelas_lain_setingkat = [k.nama for k in kelompok_kelas if k.id != kelas.id]
            if kelas_lain_setingkat:
                messages.success(
                    request,
                    f"Clustering K-Means untuk {kelas.nama} berhasil dijalankan ulang "
                    f"(digabung bersama {', '.join(kelas_lain_setingkat)} karena setingkat)."
                )
            else:
                messages.success(request, f"Clustering K-Means untuk {kelas.nama} berhasil dijalankan ulang.")
    return redirect("dashboard:cluster_overview")


@login_required
def cluster_detail(request, label):
    kelas, kelas_list = _get_kelas_aktif(request)
    siswa_list = []
    judul_kelas = kelas.nama if kelas else ""
    kelompok_kelas = []
    if kelas:
        # Sama seperti cluster_overview: gabung dengan kelas lain yang
        # setingkat (mis. Kelas 1A + Kelas 1B) supaya daftar siswanya
        # konsisten dengan ringkasan angka di halaman overview.
        kelompok_kelas = _kumpulkan_kelas_setingkat(kelas)
        judul_kelas = _judul_gabungan_kelas(kelompok_kelas)
        siswa_list = (
            Siswa.objects.filter(kelas__in=kelompok_kelas, cluster_label=label)
            .select_related("kelas")
            .order_by("kelas__nama", "rangking", "nama")
        )

    label_info = {
        "cerdas": {"nama": "Cerdas", "warna": "green",
                   "deskripsi": "Siswa dengan nilai akademik dan keaktifan tinggi secara konsisten."},
        "pintar": {"nama": "Pintar", "warna": "blue",
                   "deskripsi": "Siswa dengan performa baik namun masih ada ruang untuk berkembang."},
        "malas": {"nama": "Malas", "warna": "orange",
                  "deskripsi": "Siswa yang butuh perhatian dan dorongan tambahan dari guru."},
    }
    if label not in label_info:
        label = "cerdas"

    context = {
        "kelas": kelas,
        "kelas_list": kelas_list,
        "judul_kelas": judul_kelas,
        "tampilkan_kolom_kelas": len(kelompok_kelas) > 1,
        "siswa_list": siswa_list,
        "label": label,
        "info": label_info[label],
        "active_nav": "cluster",
    }
    return render(request, "dashboard/cluster_detail.html", context)


def _bangun_pdf_cluster(request):
    """Siapkan byte PDF hasil Cluster K-Means kelas aktif SAAT INI & nama
    file -- dipakai bersama oleh view unduh (attachment) dan preview
    (inline) supaya isinya selalu sama."""
    kelas, _ = _get_kelas_aktif(request)

    judul_kelas = ""
    siswa_snapshot = []
    ringkasan = {"cerdas": 0, "pintar": 0, "malas": 0}
    if kelas:
        kelompok_kelas = _kumpulkan_kelas_setingkat(kelas)
        judul_kelas = _judul_gabungan_kelas(kelompok_kelas)
        snapshot, ringkasan = _bangun_snapshot_cluster(kelompok_kelas, judul_kelas)
        siswa_snapshot = snapshot["siswa"]

    nama_pengunduh = request.user.first_name or request.user.username
    pdf_bytes = build_cluster_pdf(
        judul_kelas=judul_kelas,
        siswa_snapshot=siswa_snapshot,
        ringkasan=ringkasan,
        generated_by=nama_pengunduh,
    )

    nama_kelas = (judul_kelas or (kelas.nama if kelas else "kelas")).replace(" ", "-")
    nama_file = f"cluster-kmeans-{nama_kelas}-{timezone.now():%Y%m%d-%H%M}.pdf"
    return pdf_bytes, nama_file


@login_required
def cluster_pdf_view(request):
    """Unduh hasil Cluster K-Means kelas aktif saat ini sebagai PDF berkop
    surat resmi sekolah."""
    pdf_bytes, nama_file = _bangun_pdf_cluster(request)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nama_file}"'
    return response


@login_required
def cluster_pdf_preview_view(request):
    """Sama seperti cluster_pdf_view, tapi ditampilkan langsung di tab
    browser (inline) supaya bisa di-preview dulu sebelum diunduh."""
    pdf_bytes, nama_file = _bangun_pdf_cluster(request)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{nama_file}"'
    return response


def _ambil_riwayat_cluster_atau_403(request, riwayat_id):
    """Ambil RiwayatClusterKmeans sesuai hak akses: superuser bebas, guru
    cuma boleh riwayat dari kelas miliknya sendiri."""
    if request.user.is_superuser:
        return get_object_or_404(RiwayatClusterKmeans, id=riwayat_id)
    return get_object_or_404(RiwayatClusterKmeans, id=riwayat_id, kelas__guru=request.user)


@login_required
def riwayat_preview_cluster_view(request, riwayat_id):
    """Kirim hasil clustering yang tersimpan di snapshot riwayat ini sebagai
    JSON, dipakai JS untuk menampilkan tabel preview di modal (laporan.html)."""
    riwayat = _ambil_riwayat_cluster_atau_403(request, riwayat_id)
    snapshot = riwayat.snapshot or {}
    siswa_snapshot = snapshot.get("siswa", [])

    label_cluster = {"cerdas": "Cerdas", "pintar": "Pintar", "malas": "Malas"}
    baris = []
    for s in siswa_snapshot:
        baris.append([
            s.get("kelas_nama"),
            s.get("nama"),
            label_cluster.get(s.get("cluster_label"), "Belum di-cluster"),
            s.get("eksakta"),
            s.get("nonekstakta"),
            s.get("rangking"),
            s.get("skor_akademik"),
            s.get("skor_disiplin"),
        ])

    if not snapshot:
        return JsonResponse({
            "ok": False,
            "error": (
                "Riwayat ini belum punya data snapshot tersimpan (dibuat sebelum "
                "fitur preview/unduh aktif)."
            ),
        }, status=400)

    return JsonResponse({
        "ok": True,
        "judul": f"Cluster K-Means — {snapshot.get('judul_kelas') or riwayat.kelas.nama}",
        "catatan": (
            f"Hasil clustering saat dijalankan pada "
            f"{timezone.localtime(riwayat.dijalankan_pada):%d %b %Y, %H:%M} "
            f"({riwayat.get_sumber_display()})."
        ),
        "kolom": [
            "Kelas", "Nama", "Kategori Cluster", "Nilai Eksakta", "Nilai Non-Eksakta",
            "Rangking", "Skor Akademik", "Skor Disiplin",
        ],
        "baris": baris,
        "total_baris": len(baris),
        "ditampilkan": len(baris),
        "url_unduh": reverse("dashboard:riwayat_unduh_cluster", args=[riwayat.id]),
    })


@login_required
def riwayat_unduh_cluster_view(request, riwayat_id):
    """Unduh PDF hasil Cluster K-Means persis seperti saat riwayat ini
    dibuat, dibangun ulang dari snapshot yang tersimpan (bukan dari data
    siswa saat ini)."""
    riwayat = _ambil_riwayat_cluster_atau_403(request, riwayat_id)
    snapshot = riwayat.snapshot or {}

    if not snapshot:
        messages.error(request, "Riwayat ini belum punya data snapshot tersimpan untuk diunduh.")
        return redirect("dashboard:laporan")

    ringkasan = {
        "cerdas": riwayat.jumlah_cerdas,
        "pintar": riwayat.jumlah_pintar,
        "malas": riwayat.jumlah_malas,
    }
    nama_pengunduh = request.user.first_name or request.user.username
    pdf_bytes = build_cluster_pdf(
        judul_kelas=snapshot.get("judul_kelas") or riwayat.kelas.nama,
        siswa_snapshot=snapshot.get("siswa", []),
        ringkasan=ringkasan,
        generated_by=nama_pengunduh,
        tanggal=timezone.localtime(riwayat.dijalankan_pada),
    )

    nama_kelas = (snapshot.get("judul_kelas") or riwayat.kelas.nama or "kelas").replace(" ", "-")
    nama_file = f"riwayat-cluster-kmeans-{nama_kelas}-{riwayat.dijalankan_pada:%Y%m%d-%H%M}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nama_file}"'
    return response


@login_required
def naik_kelas_view(request):
    """
    Halaman 'Naik Kelas': menampilkan kelas-kelas guru ini dikelompokkan per
    tingkat (1, 2, 3, ...), supaya guru bisa menggabungkan & membagi rata
    hasil cluster dari satu tingkat (mis. 1A + 1B) ke tingkat berikutnya
    (2A + 2B).
    """
    kelas, kelas_list = _get_kelas_aktif(request)
    kelompok_tingkat_raw = get_kelas_per_tingkat(request.user)

    kelompok_tingkat = []
    for tingkat, daftar_kelas in kelompok_tingkat_raw.items():
        kelas_info = []
        total_siswa = 0
        total_belum_cluster = 0
        for k in daftar_kelas:
            jumlah = k.siswa_list.count()
            belum = k.siswa_list.filter(cluster_label__isnull=True).count()
            total_siswa += jumlah
            total_belum_cluster += belum
            kelas_info.append({"kelas": k, "jumlah": jumlah, "belum_cluster": belum})

        kelompok_tingkat.append({
            "tingkat": tingkat,
            "tingkat_baru": tingkat + 1,
            "kelas_info": kelas_info,
            "total_siswa": total_siswa,
            "total_belum_cluster": total_belum_cluster,
            "bisa_diproses": total_siswa > 0,
        })

    context = {
        "kelas": kelas,
        "kelas_list": kelas_list,
        "kelompok_tingkat": kelompok_tingkat,
        "active_nav": "cluster",
    }
    return render(request, "dashboard/naik_kelas.html", context)


@login_required
def naik_kelas_proses_view(request):
    if request.method == "POST":
        try:
            tingkat_asal = int(request.POST.get("tingkat_asal", ""))
        except (TypeError, ValueError):
            messages.error(request, "Tingkat yang dipilih tidak valid.")
            return redirect("dashboard:naik_kelas")

        try:
            hasil = proses_naik_kelas(request.user, tingkat_asal)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("dashboard:naik_kelas")

        nama_kelas_lama = ", ".join(hasil["nama_kelas_asal"])
        nama_kelas_baru = ", ".join(r["kelas"].nama for r in hasil["ringkasan_kelas_baru"])
        messages.success(
            request,
            f"{hasil['jumlah_siswa']} siswa dari {nama_kelas_lama} berhasil naik kelas dan "
            f"disebar rata (berdasarkan hasil cluster Cerdas/Pintar/Malas) ke {nama_kelas_baru}. "
            f"Kelas lama ({nama_kelas_lama}) sudah otomatis dihapus.",
        )
        if hasil["jumlah_belum_cluster"]:
            messages.warning(
                request,
                f"{hasil['jumlah_belum_cluster']} siswa belum sempat di-cluster sebelumnya, "
                "tetap disebar rata namun tanpa mempertimbangkan kategori cluster-nya.",
            )

        # arahkan kelas aktif ke salah satu kelas baru hasil proses
        if hasil["ringkasan_kelas_baru"]:
            request.session["kelas_aktif_id"] = hasil["ringkasan_kelas_baru"][0]["kelas"].id

    return redirect("dashboard:naik_kelas")


# ---------------------------------------------------------------------------
# Math Perbangku (denah pemahaman matematika per meja)
# ---------------------------------------------------------------------------
def _susun_meja_list(siswa_qs):
    """Kelompokkan queryset siswa ke meja masing-masing & hitung rata-rata +
    status pemahaman tiap meja. Dipakai bersama oleh tampilan halaman,
    generator PDF, dan snapshot riwayat supaya logikanya selalu konsisten."""
    meja_list = []
    rata_rata_kelas = None

    if siswa_qs.exists():
        rata_rata_kelas = round(siswa_qs.aggregate(avg=Avg("nilai_matematika"))["avg"], 1)

    meja_map = {}
    for s in siswa_qs:
        meja_map.setdefault(s.nomor_meja, []).append(s)

    for nomor in sorted(meja_map.keys()):
        siswa_di_meja = meja_map[nomor]
        rata = sum(s.nilai_matematika for s in siswa_di_meja) / len(siswa_di_meja)
        if rata >= 80:
            status = "paham"
        elif rata >= 60:
            status = "cukup"
        else:
            status = "kurang"
        meja_list.append({
            "nomor": nomor,
            "siswa": siswa_di_meja,
            "rata_rata": round(rata, 1),
            "status": status,
        })

    return meja_list, rata_rata_kelas


def _bangun_snapshot_perbangku(kelas):
    """Ambil kondisi meja & data siswa kelas saat ini, lalu bekukan jadi dict
    yang bisa disimpan sebagai JSON (RiwayatPerbangku.snapshot) supaya
    preview/unduh riwayat lama tetap menampilkan susunan meja & nilai siswa
    persis seperti saat itu, walau data siswa berubah/dihapus belakangan."""
    siswa_qs = kelas.siswa_list.all()
    meja_list, rata_rata_kelas = _susun_meja_list(siswa_qs)

    meja_snapshot = []
    for meja in meja_list:
        siswa_snapshot = []
        for s in meja["siswa"]:
            siswa_snapshot.append({
                "nama": s.nama,
                "eksakta": s.eksakta if s.eksakta is not None else s.nilai_matematika,
                "nonekstakta": s.nonekstakta if s.nonekstakta is not None else s.nilai_bahasa,
                "cluster_label": s.cluster_label,
            })
        meja_snapshot.append({
            "nomor": meja["nomor"],
            "rata_rata": meja["rata_rata"],
            "status": meja["status"],
            "siswa": siswa_snapshot,
        })

    return {
        "kelas_nama": kelas.nama,
        "rata_rata_kelas": rata_rata_kelas,
        "meja": meja_snapshot,
    }


@login_required
def math_perbangku(request):
    kelas, kelas_list = _get_kelas_aktif(request)
    meja_list = []
    rata_rata_kelas = None

    if kelas:
        meja_list, rata_rata_kelas = _susun_meja_list(kelas.siswa_list.all())

    context = {
        "kelas": kelas,
        "kelas_list": kelas_list,
        "meja_list": meja_list,
        "rata_rata_kelas": rata_rata_kelas,
        "active_nav": "math",
    }
    return render(request, "dashboard/math_perbangku.html", context)


@login_required
def math_perbangku_atur(request):
    """Susun ulang (atau acak ulang) posisi meja kelas aktif berdasarkan kombinasi
    cluster_label siswa (cerdas+malas, pintar+cerdas, pintar+malas)."""
    kelas, _ = _get_kelas_aktif(request)
    if request.method == "POST" and kelas:
        jumlah_meja = atur_meja_per_kelas(kelas)
        if jumlah_meja:
            snapshot = _bangun_snapshot_perbangku(kelas)
            RiwayatPerbangku.objects.create(
                kelas=kelas,
                diatur_oleh=request.user,
                jumlah_meja=jumlah_meja,
                jumlah_siswa=kelas.siswa_list.count(),
                snapshot=snapshot,
            )
            messages.success(request, f"Posisi meja {kelas.nama} berhasil diatur ulang ({jumlah_meja} meja).")
        else:
            messages.warning(request, "Belum ada siswa di kelas ini untuk diatur posisinya.")
    return redirect("dashboard:math_perbangku")


def _bangun_pdf_math_perbangku(request):
    """Siapkan byte PDF Math Perbangku & nama file -- dipakai bersama oleh
    view unduh (attachment) dan preview (inline) supaya isinya selalu sama."""
    kelas, _ = _get_kelas_aktif(request)

    meja_list = []
    rata_rata_kelas = None
    if kelas:
        meja_list, rata_rata_kelas = _susun_meja_list(kelas.siswa_list.all())

    nama_pengunduh = request.user.first_name or request.user.username
    pdf_bytes = build_perbangku_pdf(
        kelas=kelas,
        meja_list=meja_list,
        rata_rata_kelas=rata_rata_kelas,
        generated_by=nama_pengunduh,
    )

    nama_kelas = kelas.nama if kelas else "kelas"
    nama_file = f"math-perbangku-{nama_kelas}-{timezone.now():%Y%m%d-%H%M}.pdf".replace(" ", "-")
    return pdf_bytes, nama_file


@login_required
def math_perbangku_pdf_view(request):
    """Unduh hasil pengaturan meja (Math Perbangku) kelas aktif sebagai PDF
    berkop surat resmi sekolah, lengkap dengan tempat/tanggal dan tanda
    tangan Kepala Sekolah di pojok kanan bawah."""
    pdf_bytes, nama_file = _bangun_pdf_math_perbangku(request)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nama_file}"'
    return response


@login_required
def math_perbangku_pdf_preview_view(request):
    """Sama seperti math_perbangku_pdf_view, tapi ditampilkan langsung di
    tab browser (inline) supaya bisa di-preview dulu sebelum diunduh."""
    pdf_bytes, nama_file = _bangun_pdf_math_perbangku(request)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{nama_file}"'
    return response


def _ambil_riwayat_perbangku_atau_403(request, riwayat_id):
    """Ambil RiwayatPerbangku sesuai hak akses: superuser bebas, guru cuma
    boleh riwayat dari kelas miliknya sendiri."""
    if request.user.is_superuser:
        return get_object_or_404(RiwayatPerbangku, id=riwayat_id)
    return get_object_or_404(RiwayatPerbangku, id=riwayat_id, kelas__guru=request.user)


@login_required
def riwayat_preview_perbangku_view(request, riwayat_id):
    """Kirim susunan meja & data siswa yang tersimpan di snapshot riwayat ini
    sebagai JSON, dipakai JS untuk menampilkan tabel preview di modal."""
    riwayat = _ambil_riwayat_perbangku_atau_403(request, riwayat_id)
    snapshot = riwayat.snapshot or {}
    meja_snapshot = snapshot.get("meja", [])

    label_cluster = {"cerdas": "Cerdas", "pintar": "Pintar", "malas": "Malas"}
    baris = []
    for meja in meja_snapshot:
        for s in meja.get("siswa", []):
            baris.append([
                meja.get("nomor"),
                s.get("nama"),
                label_cluster.get(s.get("cluster_label"), "Belum di-cluster"),
                s.get("eksakta"),
                s.get("nonekstakta"),
            ])

    if not snapshot:
        return JsonResponse({
            "ok": False,
            "error": (
                "Riwayat ini belum punya data snapshot tersimpan (dibuat sebelum "
                "fitur preview/unduh aktif)."
            ),
        }, status=400)

    return JsonResponse({
        "ok": True,
        "judul": f"Perbangku — {riwayat.kelas.nama}",
        "catatan": (
            f"Susunan meja saat diatur pada "
            f"{timezone.localtime(riwayat.diatur_pada):%d %b %Y, %H:%M}."
        ),
        "kolom": ["Meja", "Nama", "Kategori Cluster", "Nilai Eksakta", "Nilai Non-Eksakta"],
        "baris": baris,
        "total_baris": len(baris),
        "ditampilkan": len(baris),
        "url_unduh": reverse("dashboard:riwayat_unduh_perbangku", args=[riwayat.id]),
    })


@login_required
def riwayat_unduh_perbangku_view(request, riwayat_id):
    """Unduh PDF susunan meja persis seperti saat riwayat ini dibuat, dibangun
    ulang dari snapshot yang tersimpan (bukan dari data siswa saat ini)."""
    from types import SimpleNamespace

    riwayat = _ambil_riwayat_perbangku_atau_403(request, riwayat_id)
    snapshot = riwayat.snapshot or {}
    meja_snapshot = snapshot.get("meja", [])

    if not snapshot:
        messages.error(request, "Riwayat ini belum punya data snapshot tersimpan untuk diunduh.")
        return redirect("dashboard:laporan")

    kelas_ns = SimpleNamespace(nama=snapshot.get("kelas_nama") or riwayat.kelas.nama)
    meja_list = []
    for meja in meja_snapshot:
        siswa_ns = [
            SimpleNamespace(
                nama=s.get("nama"),
                eksakta=s.get("eksakta"),
                nonekstakta=s.get("nonekstakta"),
                nilai_matematika=s.get("eksakta"),
                nilai_bahasa=s.get("nonekstakta"),
                cluster_label=s.get("cluster_label"),
                kelas=kelas_ns,
            )
            for s in meja.get("siswa", [])
        ]
        meja_list.append({
            "nomor": meja.get("nomor"),
            "siswa": siswa_ns,
            "rata_rata": meja.get("rata_rata"),
            "status": meja.get("status"),
        })

    nama_pengunduh = request.user.first_name or request.user.username
    pdf_bytes = build_perbangku_pdf(
        kelas=kelas_ns,
        meja_list=meja_list,
        rata_rata_kelas=snapshot.get("rata_rata_kelas"),
        generated_by=nama_pengunduh,
        tanggal=timezone.localtime(riwayat.diatur_pada),
    )

    nama_kelas = (snapshot.get("kelas_nama") or riwayat.kelas.nama or "kelas").replace(" ", "-")
    nama_file = f"riwayat-perbangku-{nama_kelas}-{riwayat.diatur_pada:%Y%m%d-%H%M}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nama_file}"'
    return response


# ---------------------------------------------------------------------------
# Upload File
# ---------------------------------------------------------------------------
@login_required
def upload_file_view(request):
    kelas, kelas_list = _get_kelas_aktif(request)

    if request.user.is_superuser:
        riwayat_upload = RiwayatUploadXlsx.objects.select_related("kelas", "diupload_oleh").all()[:30]
    elif kelas:
        riwayat_upload = kelas.riwayat_upload.select_related("diupload_oleh").all()[:15]
    else:
        riwayat_upload = RiwayatUploadXlsx.objects.none()

    context = {
        "kelas": kelas,
        "kelas_list": kelas_list,
        "siswa_form": SiswaManualForm(),
        "upload_form": UploadDataSiswaForm(user=request.user),
        "riwayat_upload": riwayat_upload,
        "active_nav": "upload",
    }
    return render(request, "dashboard/upload_file.html", context)


@login_required
def tambah_siswa_manual_view(request):
    """
    Tambah 1 siswa secara manual ke kelas aktif -- alternatif kalau upload
    file Excel leger di halaman Upload tidak berhasil/format filenya tidak
    sesuai. Begitu siswa baru tersimpan, clustering K-Means langsung
    dijalankan ulang untuk seluruh kelas tsb (run_clustering, lihat
    dashboard/clustering.py) supaya hasil Cerdas/Pintar/Malas selalu sinkron.
    """
    kelas, _ = _get_kelas_aktif(request)

    if request.method == "POST" and kelas:
        form = SiswaManualForm(request.POST)
        if form.is_valid():
            siswa = form.save(commit=False)
            siswa.kelas = kelas
            if not siswa.nomor_meja:
                siswa.nomor_meja = kelas.siswa_list.count() + 1

            # Selaraskan ke field lama (nilai_matematika/nilai_bahasa/keaktifan/
            # kehadiran_persen) yang masih dipakai fitur lain (Math Perbangku,
            # Laporan) -- dihitung dari Eksakta/NonEksakta/Izin/Sakit/Alpa
            # dengan rumus yang sama seperti excel_import.py.
            izin, sakit, alpa = siswa.izin or 0.0, siswa.sakit or 0.0, siswa.alpa or 0.0
            siswa.nilai_matematika = siswa.eksakta
            siswa.nilai_bahasa = siswa.nonekstakta
            siswa.keaktifan = siswa.nonekstakta
            siswa.kehadiran_persen = round(max(0.0, min(100.0, 100 - (izin + sakit + alpa * 3))), 1)

            siswa.save()
            run_clustering(kelas)
            messages.success(
                request,
                f"{siswa.nama} berhasil ditambahkan secara manual ke {kelas.nama}, "
                "dan clustering K-Means kelas ini sudah dijalankan ulang.",
            )
        else:
            for field_errors in form.errors.values():
                for err in field_errors:
                    messages.error(request, err)

    return redirect("dashboard:upload_file")


@login_required
def materi_hapus_view(request, materi_id):
    materi = get_object_or_404(MateriUpload, id=materi_id, kelas__guru=request.user)
    materi.delete()
    messages.success(request, "File dihapus.")
    return redirect("dashboard:upload_file")


# ---------------------------------------------------------------------------
# Riwayat Laporan (riwayat aktivitas guru: upload Excel, login, dll)
# ---------------------------------------------------------------------------
def _kumpulkan_data_laporan(request, batas_riwayat=None):
    """
    Kumpulkan seluruh data untuk halaman Riwayat Laporan.
    Dipakai bersama oleh tampilan HTML (laporan_view) dan unduhan PDF
    (laporan_pdf_view) supaya datanya selalu konsisten.

    batas_riwayat: jumlah maksimum baris riwayat upload yang diambil.
    None berarti tanpa batas (dipakai untuk PDF supaya laporannya lengkap).
    """
    kelas, kelas_list = _get_kelas_aktif(request)

    guru_stats = None
    total_guru = None
    total_upload_excel = None

    if request.user.is_superuser:
        # Superuser bisa memantau seluruh guru: berapa kali tiap guru upload
        # Excel, kapan terakhir login, dan sejak kapan bergabung.
        guru_qs = User.objects.filter(is_superuser=False).order_by(
            "first_name", "username"
        )
        guru_stats = []
        for g in guru_qs:
            guru_stats.append({
                "user": g,
                "jumlah_kelas": Kelas.objects.filter(guru=g).count(),
                "jumlah_siswa": Siswa.objects.filter(kelas__guru=g).count(),
                "jumlah_upload": RiwayatUploadXlsx.objects.filter(kelas__guru=g).count(),
                "terakhir_login": g.last_login,
                "bergabung_sejak": g.date_joined,
            })

        riwayat_upload = (
            RiwayatUploadXlsx.objects
            .select_related("kelas", "kelas__guru", "diupload_oleh")
            .all()
        )
        riwayat_perbangku = (
            RiwayatPerbangku.objects
            .select_related("kelas", "kelas__guru", "diatur_oleh")
            .all()
        )
        riwayat_cluster = (
            RiwayatClusterKmeans.objects
            .select_related("kelas", "kelas__guru", "dijalankan_oleh")
            .all()
        )
        if batas_riwayat is not None:
            riwayat_upload = riwayat_upload[:batas_riwayat]
            riwayat_perbangku = riwayat_perbangku[:batas_riwayat]
            riwayat_cluster = riwayat_cluster[:batas_riwayat]
        total_guru = guru_qs.count()
        total_upload_excel = RiwayatUploadXlsx.objects.count()
    elif kelas:
        # Guru biasa hanya melihat riwayat upload untuk kelas-kelas miliknya sendiri.
        riwayat_upload = (
            RiwayatUploadXlsx.objects
            .filter(kelas__guru=request.user)
            .select_related("kelas", "diupload_oleh")
            .all()
        )
        riwayat_perbangku = (
            RiwayatPerbangku.objects
            .filter(kelas__guru=request.user)
            .select_related("kelas", "diatur_oleh")
            .all()
        )
        riwayat_cluster = (
            RiwayatClusterKmeans.objects
            .filter(kelas__guru=request.user)
            .select_related("kelas", "dijalankan_oleh")
            .all()
        )
        if batas_riwayat is not None:
            riwayat_upload = riwayat_upload[:batas_riwayat]
            riwayat_perbangku = riwayat_perbangku[:batas_riwayat]
            riwayat_cluster = riwayat_cluster[:batas_riwayat]
        total_upload_excel = RiwayatUploadXlsx.objects.filter(kelas__guru=request.user).count()
    else:
        riwayat_upload = RiwayatUploadXlsx.objects.none()
        riwayat_perbangku = RiwayatPerbangku.objects.none()
        riwayat_cluster = RiwayatClusterKmeans.objects.none()

    return {
        "kelas": kelas,
        "kelas_list": kelas_list,
        "guru_stats": guru_stats,
        "riwayat_upload": riwayat_upload,
        "riwayat_perbangku": riwayat_perbangku,
        "riwayat_cluster": riwayat_cluster,
        "total_guru": total_guru,
        "total_upload_excel": total_upload_excel,
        "terakhir_login_saya": request.user.last_login,
        "bergabung_sejak_saya": request.user.date_joined,
    }


@login_required
def laporan_view(request):
    context = _kumpulkan_data_laporan(request, batas_riwayat=100 if request.user.is_superuser else 50)
    context["active_nav"] = "laporan"
    return render(request, "dashboard/laporan.html", context)


def _ambil_riwayat_atau_403(request, riwayat_id):
    """Ambil RiwayatUploadXlsx sesuai hak akses: superuser bebas, guru cuma
    boleh riwayat dari kelas miliknya sendiri."""
    if request.user.is_superuser:
        return get_object_or_404(RiwayatUploadXlsx, id=riwayat_id)
    return get_object_or_404(RiwayatUploadXlsx, id=riwayat_id, kelas__guru=request.user)


@login_required
def riwayat_unduh_asli_view(request, riwayat_id):
    """
    Unduh file Excel ASLI yang diupload guru (belum diolah).

    Untuk upload lama (sebelum file fisiknya mulai disimpan), file asli
    memang sudah tidak ada di server -- jadi sebagai gantinya kita bikinkan
    file Excel pengganti dari data siswa yang tersimpan di database, supaya
    tombol Unduh tetap selalu berfungsi.
    """
    riwayat = _ambil_riwayat_atau_403(request, riwayat_id)

    if riwayat.file_asli:
        response = HttpResponse(
            riwayat.file_asli.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        nama = riwayat.nama_file or "data_asli.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{nama}"'
        return response

    # --- Fallback untuk riwayat lama tanpa file fisik tersimpan ---
    # Struktur tabelnya dibuat SAMA seperti leger asli (grup kolom NO/NAMA/
    # NISN/NIS, MATA PELAJARAN, Ketidakhadiran, Ekstra Kurikuler) -- BUKAN
    # kolom hasil olahan (itu sudah ada di section "Olah Data" sendiri).
    # Nilai per-mapel & per-kegiatan ekskul memang sudah tidak tersimpan di
    # database (yang tersimpan cuma hasil rata-ratanya), jadi sel-sel itu
    # dikosongkan ("-") apa adanya -- bukan diisi angka olahan supaya gak
    # ketuker sama data mentah asli.
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    siswa_list = riwayat.kelas.siswa_list.all().order_by("rangking", "nama")

    KOLOM_MAPEL = ["PAIDBP", "PP", "BI", "MU", "IPADSI", "PJODK", "BI", "SR", "PLBJ", "GKS"]
    KOLOM_KETIDAKHADIRAN = ["Sakit", "Izin", "Alpa"]
    KOLOM_EKSKUL = ["Marawis", "Menari", "Pramuka Siaga", "Melukis"]

    # Urutan & lebar grup: NO, NAMA SISWA, NISN, NIS, [MATA PELAJARAN...],
    # [Ketidakhadiran...], [Ekstra Kurikuler...]
    kolom_awal = ["NO", "NAMA SISWA", "NISN", "NIS"]
    total_kolom = len(kolom_awal) + len(KOLOM_MAPEL) + len(KOLOM_KETIDAKHADIRAN) + len(KOLOM_EKSKUL)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Asli"

    tebal = Font(bold=True, color="FFFFFF")
    isi_biru = openpyxl.styles.PatternFill("solid", fgColor="4F6FA5")
    tengah = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_tipis = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

    # Baris info + catatan keterbatasan data
    ws.append([f"LEGER NILAI — {riwayat.kelas.nama} (rekonstruksi dari data tersimpan)"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_kolom)
    ws.append([
        f"File fisik asli '{riwayat.nama_file}' tidak tersimpan di server (upload sebelum fitur "
        f"ini aktif). Kolom bertanda \"-\" berarti nilai mentahnya sudah tidak tersedia -- hanya "
        f"struktur tabel & data yang masih ada di database (nama, NISN, NIS, ketidakhadiran) yang "
        f"ditampilkan. Untuk nilai hasil olahan, lihat bagian \"Olah Data\"."
    ])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_kolom)
    ws["A1"].font = Font(bold=True, size=11)
    ws["A2"].font = Font(italic=True, size=9, color="888888")
    ws["A2"].alignment = Alignment(wrap_text=True)

    # Baris header grup (mis. "MATA PELAJARAN" merentang 10 kolom)
    baris_grup = 3
    baris_sub = 4
    ws.cell(row=baris_grup, column=1, value="NO")
    ws.cell(row=baris_grup, column=2, value="NAMA SISWA")
    ws.cell(row=baris_grup, column=3, value="NISN")
    ws.cell(row=baris_grup, column=4, value="NIS")
    for c in range(1, 5):
        ws.merge_cells(start_row=baris_grup, start_column=c, end_row=baris_sub, end_column=c)

    kolom_mulai = 5
    ws.cell(row=baris_grup, column=kolom_mulai, value="MATA PELAJARAN")
    ws.merge_cells(start_row=baris_grup, start_column=kolom_mulai,
                   end_row=baris_grup, end_column=kolom_mulai + len(KOLOM_MAPEL) - 1)
    for i, nama_kol in enumerate(KOLOM_MAPEL):
        ws.cell(row=baris_sub, column=kolom_mulai + i, value=nama_kol)

    kolom_ketidakhadiran = kolom_mulai + len(KOLOM_MAPEL)
    ws.cell(row=baris_grup, column=kolom_ketidakhadiran, value="Ketidakhadiran")
    ws.merge_cells(start_row=baris_grup, start_column=kolom_ketidakhadiran,
                   end_row=baris_grup, end_column=kolom_ketidakhadiran + len(KOLOM_KETIDAKHADIRAN) - 1)
    for i, nama_kol in enumerate(KOLOM_KETIDAKHADIRAN):
        ws.cell(row=baris_sub, column=kolom_ketidakhadiran + i, value=nama_kol)

    kolom_ekskul = kolom_ketidakhadiran + len(KOLOM_KETIDAKHADIRAN)
    ws.cell(row=baris_grup, column=kolom_ekskul, value="Ekstra Kurikuler")
    ws.merge_cells(start_row=baris_grup, start_column=kolom_ekskul,
                   end_row=baris_grup, end_column=kolom_ekskul + len(KOLOM_EKSKUL) - 1)
    for i, nama_kol in enumerate(KOLOM_EKSKUL):
        ws.cell(row=baris_sub, column=kolom_ekskul + i, value=nama_kol)

    for row in ws.iter_rows(min_row=baris_grup, max_row=baris_sub, max_col=total_kolom):
        for cell in row:
            cell.font = tebal
            cell.fill = isi_biru
            cell.alignment = tengah
            cell.border = border_tipis

    # Baris data: NO/NAMA/NISN/NIS/Ketidakhadiran terisi dari database,
    # MATA PELAJARAN & rincian Ekstra Kurikuler dikosongkan ("-") karena
    # nilainya sudah tidak tersimpan mentah.
    baris_ke = baris_sub + 1
    for i, s in enumerate(siswa_list, start=1):
        ws.cell(row=baris_ke, column=1, value=i)
        ws.cell(row=baris_ke, column=2, value=s.nama)
        ws.cell(row=baris_ke, column=3, value=s.nisn or "-")
        ws.cell(row=baris_ke, column=4, value=s.nis or "-")
        for offset in range(len(KOLOM_MAPEL)):
            ws.cell(row=baris_ke, column=kolom_mulai + offset, value="-")
        ws.cell(row=baris_ke, column=kolom_ketidakhadiran, value=s.sakit or 0)
        ws.cell(row=baris_ke, column=kolom_ketidakhadiran + 1, value=s.izin or 0)
        ws.cell(row=baris_ke, column=kolom_ketidakhadiran + 2, value=s.alpa or 0)
        for offset in range(len(KOLOM_EKSKUL)):
            ws.cell(row=baris_ke, column=kolom_ekskul + offset, value="-")
        baris_ke += 1

    for col_idx in range(1, total_kolom + 1):
        panjang = max(
            (len(str(ws.cell(row=r, column=col_idx).value))
             for r in range(baris_sub, ws.max_row + 1)
             if ws.cell(row=r, column=col_idx).value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max(9, min(panjang + 2, 40))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    nama_dasar = (riwayat.nama_file or "data_asli.xlsx").rsplit(".", 1)[0]
    response["Content-Disposition"] = f'attachment; filename="{nama_dasar}_rekonstruksi.xlsx"'
    wb.save(response)
    return response


@login_required
def riwayat_unduh_olahan_view(request, riwayat_id):
    """
    Unduh data HASIL OLAHAN (setelah diproses & dilatih K-Means) untuk kelas
    dari riwayat upload ini, sebagai PDF berkop surat resmi sekolah (sama
    seperti laporan Math Perbangku & Cluster K-Means) dengan kolom:
    No, Nama, Nilai Eksakta & Non-Eksakta, Izin, Sakit, Alpa.
    """
    riwayat = _ambil_riwayat_atau_403(request, riwayat_id)
    siswa_qs = riwayat.kelas.siswa_list.all().order_by("rangking", "nama")

    siswa_list = [
        {
            "nama": s.nama,
            "eksakta": s.eksakta if s.eksakta is not None else s.nilai_matematika,
            "nonekstakta": s.nonekstakta if s.nonekstakta is not None else s.nilai_bahasa,
            "izin": s.izin or 0,
            "sakit": s.sakit or 0,
            "alpa": s.alpa or 0,
        }
        for s in siswa_qs
    ]

    nama_pengunduh = request.user.first_name or request.user.username
    pdf_bytes = build_data_olahan_pdf(
        kelas_nama=riwayat.kelas.nama,
        siswa_list=siswa_list,
        generated_by=nama_pengunduh,
        nama_file_asli=riwayat.nama_file,
    )

    nama_kelas = (riwayat.kelas.nama or "kelas").replace(" ", "-")
    nama_file = f"data-olahan-{nama_kelas}-{riwayat.diupload_pada:%Y%m%d-%H%M}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nama_file}"'
    return response


# Batas jumlah baris yang ditampilkan di modal preview, supaya modal tidak
# kebanyakan data kalau kelasnya besar / file excel-nya banyak baris kosong.
_BATAS_BARIS_PREVIEW = 60


@login_required
def riwayat_preview_olahan_view(request, riwayat_id):
    """
    Kirim data HASIL OLAHAN (sama persis dengan yang dipakai di
    riwayat_unduh_olahan_view) sebagai JSON, dipakai JS untuk menampilkan
    tabel preview di dalam modal sebelum diunduh.
    """
    riwayat = _ambil_riwayat_atau_403(request, riwayat_id)
    siswa_list = riwayat.kelas.siswa_list.all().order_by("rangking", "nama")

    # Header 2 baris: "Nilai" digabung (colspan 2) membawahi Eksakta &
    # Non-Eksakta, kolom lain digabung vertikal (rowspan 2).
    header_grup = [
        {"label": "No", "rowspan": 2},
        {"label": "Nama", "rowspan": 2},
        {"label": "Nilai", "colspan": 2},
        {"label": "Izin", "rowspan": 2},
        {"label": "Sakit", "rowspan": 2},
        {"label": "Alpa", "rowspan": 2},
    ]
    header_sub = ["Eksakta", "Non-Eksakta"]

    total_baris = siswa_list.count()
    baris = []
    for i, s in enumerate(siswa_list[:_BATAS_BARIS_PREVIEW], start=1):
        baris.append([
            i,
            s.nama,
            s.eksakta if s.eksakta is not None else s.nilai_matematika,
            s.nonekstakta if s.nonekstakta is not None else s.nilai_bahasa,
            s.izin or 0,
            s.sakit or 0,
            s.alpa or 0,
        ])

    return JsonResponse({
        "ok": True,
        "judul": f"Data Olahan — {riwayat.kelas.nama}",
        "catatan": "Data hasil olahan (setelah diproses & dilatih K-Means).",
        "kolom": ["No", "Nama", "Nilai Eksakta", "Nilai Non-Eksakta", "Izin", "Sakit", "Alpa"],
        "header_grup": header_grup,
        "header_sub": header_sub,
        "baris": baris,
        "total_baris": total_baris,
        "ditampilkan": len(baris),
        "url_unduh": reverse("dashboard:riwayat_unduh_olahan", args=[riwayat.id]),
    })


@login_required
def riwayat_preview_asli_view(request, riwayat_id):
    """
    Kirim isi file Excel ASLI (atau rekonstruksinya kalau file fisiknya
    sudah tidak ada, sama seperti di riwayat_unduh_asli_view) sebagai JSON
    untuk ditampilkan di modal preview.
    """
    import openpyxl

    riwayat = _ambil_riwayat_atau_403(request, riwayat_id)

    if riwayat.file_asli:
        # Baca langsung isi file excel asli yang tersimpan di server.
        try:
            wb = openpyxl.load_workbook(riwayat.file_asli, data_only=True)
            ws = wb.active
        except Exception:
            return JsonResponse({
                "ok": False,
                "error": "File Excel asli tidak bisa dibaca (mungkin rusak/format tidak didukung).",
            }, status=400)

        semua_baris = list(ws.iter_rows(values_only=True))
        total_baris = len(semua_baris)
        dipotong = semua_baris[:_BATAS_BARIS_PREVIEW]

        # Cari lebar kolom terpanjang di antara baris yang ditampilkan supaya
        # tabelnya rapi (sebagian baris judul biasanya lebih pendek dari data).
        lebar = max((len(r) for r in dipotong), default=0)
        baris = [
            ["" if v is None else v for v in (list(r) + [None] * (lebar - len(r)))]
            for r in dipotong
        ]

        return JsonResponse({
            "ok": True,
            "judul": f"Data Asli — {riwayat.nama_file}",
            "catatan": "Isi apa adanya dari file Excel yang diupload (baris & kolom mengikuti file aslinya).",
            "kolom": None,  # tidak ada header baku, tampilkan sebagai grid polos
            "baris": baris,
            "total_baris": total_baris,
            "ditampilkan": len(baris),
            "url_unduh": reverse("dashboard:riwayat_unduh_asli", args=[riwayat.id]),
        })

    # --- Fallback untuk riwayat lama tanpa file fisik tersimpan ---
    # Struktur sama seperti rekonstruksi di riwayat_unduh_asli_view, tapi
    # dikirim sebagai JSON supaya bisa dirender jadi tabel di modal.
    siswa_list = riwayat.kelas.siswa_list.all().order_by("rangking", "nama")

    kolom = [
        "NO", "NAMA SISWA", "NISN", "NIS",
        "Sakit", "Izin", "Alpa",
    ]
    total_baris = siswa_list.count()
    baris = []
    for i, s in enumerate(siswa_list[:_BATAS_BARIS_PREVIEW], start=1):
        baris.append([
            i, s.nama, s.nisn or "-", s.nis or "-",
            s.sakit or 0, s.izin or 0, s.alpa or 0,
        ])

    return JsonResponse({
        "ok": True,
        "judul": f"Data Asli — {riwayat.nama_file} (rekonstruksi)",
        "catatan": (
            "File fisik asli tidak tersimpan di server (upload sebelum fitur ini aktif). "
            "Kolom nilai per-mapel & rincian ekstrakurikuler sudah tidak tersedia — hanya "
            "data yang masih ada di database yang ditampilkan."
        ),
        "kolom": kolom,
        "baris": baris,
        "total_baris": total_baris,
        "ditampilkan": len(baris),
        "url_unduh": reverse("dashboard:riwayat_unduh_asli", args=[riwayat.id]),
    })


# ---------------------------------------------------------------------------
# Avatar Guru (upload via kamera atau file)
# ---------------------------------------------------------------------------
@login_required
def upload_foto_guru(request):
    if request.method == "POST" and request.FILES.get("foto"):
        profil, _ = ProfilGuru.objects.get_or_create(user=request.user)

        # Simpan path foto LAMA dulu (kalau ada), tapi JANGAN dihapus sekarang.
        old_path = profil.foto.path if profil.foto else None

        # Simpan foto BARU dulu -- kalau nama filenya bentrok dengan yang lama
        # (mis. JS selalu ngirim nama file yang sama, "foto-profil.jpg"), storage
        # Django otomatis kasih akhiran unik (foto-profil_AbC123.jpg) supaya URL-nya
        # selalu baru. Ini penting: kalau URL-nya sama persis seperti sebelumnya,
        # browser bakal nampilin foto LAMA dari cache walau filenya sudah diganti.
        profil.foto = request.FILES["foto"]
        profil.save()

        # Baru sekarang aman hapus file lama dari disk (setelah file baru
        # tersimpan dengan nama yang sudah pasti berbeda).
        if old_path and os.path.exists(old_path) and old_path != profil.foto.path:
            os.remove(old_path)

        from django.http import JsonResponse
        return JsonResponse({"ok": True, "url": profil.foto.url})
    from django.http import JsonResponse
    return JsonResponse({"ok": False}, status=400)


# ---------------------------------------------------------------------------
# Pengaturan
# ---------------------------------------------------------------------------
@login_required
def pengaturan_view(request):
    kelas, kelas_list = _get_kelas_aktif(request)
    profil, _ = ProfilGuru.objects.get_or_create(user=request.user)

    profil_form = ProfilGuruForm(instance=profil)
    password_form = _style_password_form(PasswordChangeForm(user=request.user))

    if request.method == "POST":
        if "simpan_profil" in request.POST:
            profil_form = ProfilGuruForm(request.POST, request.FILES, instance=profil)
            if profil_form.is_valid():
                profil_form.save()
                messages.success(request, "Profil berhasil disimpan.")
                return redirect("dashboard:pengaturan")
        elif "ganti_password" in request.POST:
            password_form = _style_password_form(PasswordChangeForm(user=request.user, data=request.POST))
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Password berhasil diubah.")
                return redirect("dashboard:pengaturan")

    context = {
        "kelas": kelas,
        "kelas_list": kelas_list,
        "profil_form": profil_form,
        "password_form": password_form,
        "active_nav": "pengaturan",
    }
    return render(request, "dashboard/pengaturan.html", context)


@login_required
def hapus_akun_view(request):
    if request.method == "POST":
        password = request.POST.get("password", "")
        user = authenticate(request, username=request.user.username, password=password)
        if user is None:
            messages.error(request, "Password salah. Akun tidak dihapus.")
            return redirect("dashboard:pengaturan")

        logout(request)
        user.delete()
        messages.success(request, "Akun Anda telah dihapus secara permanen.")
        return redirect("dashboard:login")

    return redirect("dashboard:pengaturan")


# ===================== PWA =====================
@cache_control(max_age=0, no_cache=True, no_store=True, must_revalidate=True)
def service_worker(request):
    """Serve the service worker from root scope so it can control all pages."""
    import os
    from django.conf import settings
    sw_path = os.path.join(settings.STATIC_ROOT or settings.BASE_DIR / 'static', 'js', 'sw.js')
    # Fallback: read from staticfiles directory
    if not os.path.exists(sw_path):
        sw_path = os.path.join(settings.BASE_DIR, 'static', 'js', 'sw.js')
    try:
        with open(sw_path, 'r') as f:
            content = f.read()
    except FileNotFoundError:
        content = '// Service Worker not found'
    return HttpResponse(content, content_type='application/javascript')
